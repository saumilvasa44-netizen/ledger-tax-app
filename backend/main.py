import io
import re
import math
import asyncio
from datetime import datetime, date

import pdfplumber
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = FastAPI(title="ITR Tax Tracker API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # tighten to your deployed frontend URL before wide sharing
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------------
# MODELS
# ---------------------------------------------------------------------------

class TaxInput(BaseModel):
    fy: str  # "FY 2025-26" or "FY 2024-25"
    age_group: str = "Below 60"  # "Below 60" | "60 to 79" | "80 and above"

    gross_salary: float = 0.0
    nps_employer: float = 0.0  # 80CCD(2) - allowed in both regimes

    # Other income (both regimes)
    sb_interest: float = 0.0
    fd_interest: float = 0.0
    dividend_income: float = 0.0

    # Old-regime-only exemptions
    hra_exemption: float = 0.0
    lta_exemption: float = 0.0
    other_exemptions: float = 0.0

    # Old-regime-only Chapter VI-A deductions
    ded_80c: float = 0.0
    ded_80ccd1b: float = 0.0
    ded_80d: float = 0.0
    ded_80tta_ttb: float = 0.0
    ded_other: float = 0.0

    # Tax already paid (both regimes)
    tds_salary: float = 0.0
    tds_other: float = 0.0
    advance_tax: float = 0.0


class RegimeResult(BaseModel):
    taxable_income: int
    base_tax: int
    tax_with_cess: int
    net_payable: int
    is_refund: bool


class TaxResult(BaseModel):
    new_regime: RegimeResult
    old_regime: RegimeResult
    total_tds_paid: int
    better_regime: str  # "new" | "old" | "same"


class ExtractedFields(BaseModel):
    doc_type: str
    gross_salary: int | None = None
    tds: int | None = None
    sb_interest: int | None = None
    fd_interest: int | None = None
    dividend_income: int | None = None
    raw_text_preview: str = ""


# ---------------------------------------------------------------------------
# TAX CALCULATION
# ---------------------------------------------------------------------------

STANDARD_DEDUCTION_NEW = 75000
STANDARD_DEDUCTION_OLD = 50000
SECTION_208_THRESHOLD = 10000  # no advance-tax obligation (hence no 234B/234C) below this


def round_10(value: float) -> int:
    """Section 288A / 288B style rounding: round to the nearest multiple of
    ten rupees, with the .5 case always rounding away from zero (never
    banker's rounding) - matches how ITR computation software rounds."""
    sign = -1 if value < 0 else 1
    v = abs(value)
    return int(sign * (math.floor(v / 10 + 0.5) * 10))


def round_half_up(value: float) -> int:
    sign = -1 if value < 0 else 1
    v = abs(value)
    return int(sign * math.floor(v + 0.5))


def floor_100(value: float) -> int:
    """The 'Remaining Tax Due (Round off in 100 Rs.)' step used before
    applying the 234B/234C monthly rate - always rounds DOWN to the nearest
    hundred, never to the nearest hundred."""
    if value <= 0:
        return 0
    return int(math.floor(value / 100) * 100)


def new_regime_brackets(fy: str):
    if "2025-26" in fy:
        return (
            [
                (0, 400000, 0.00), (400000, 800000, 0.05), (800000, 1200000, 0.10),
                (1200000, 1600000, 0.15), (1600000, 2000000, 0.20), (2000000, 2400000, 0.25),
                (2400000, None, 0.30),
            ],
            1200000,
        )
    return (
        [
            (0, 300000, 0.00), (300000, 600000, 0.05), (600000, 900000, 0.10),
            (900000, 1200000, 0.15), (1200000, 1500000, 0.20), (1500000, None, 0.30),
        ],
        700000,
    )


def old_regime_brackets(age_group: str):
    if age_group == "60 to 79":
        exempt_limit = 300000
    elif age_group == "80 and above":
        exempt_limit = 500000
    else:
        exempt_limit = 250000
    return [
        (0, exempt_limit, 0.00),
        (exempt_limit, 500000, 0.05),
        (500000, 1000000, 0.20),
        (1000000, None, 0.30),
    ], exempt_limit


def compute_slab_tax(taxable_income: float, brackets: list):
    """Walks each slab bracket and returns (tax, breakdown) where breakdown
    is a list of line items suitable for a 'Tax calculation on Normal
    income' style display, mirroring how ITR computation software shows it."""
    tax = 0.0
    breakdown = []
    for lo, hi, rate in brackets:
        if taxable_income <= lo:
            break
        upper = hi if hi is not None else taxable_income
        amount_in_band = min(taxable_income, upper) - lo
        if amount_in_band <= 0:
            continue
        band_tax = amount_in_band * rate
        tax += band_tax
        breakdown.append({
            "from": round(lo),
            "to": round(min(taxable_income, upper)),
            "rate_pct": round(rate * 100, 2),
            "taxable_amount": round(amount_in_band),
            "tax": round(band_tax),
        })
    return tax, breakdown


def compute_new_regime_tax(taxable_income: float, fy: str):
    brackets, rebate_limit = new_regime_brackets(fy)
    tax, breakdown = compute_slab_tax(taxable_income, brackets)
    if taxable_income <= rebate_limit:
        tax = 0.0
        breakdown = []
    return tax, breakdown


def compute_old_regime_tax(taxable_income: float, age_group: str):
    brackets, _ = old_regime_brackets(age_group)
    tax, breakdown = compute_slab_tax(taxable_income, brackets)
    if taxable_income <= 500000:
        tax = 0.0
        breakdown = []
    return tax, breakdown


def build_regime_result(taxable_income: float, base_tax: float, total_tds_paid: float) -> RegimeResult:
    tax_with_cess = round_half_up(base_tax * 1.04)
    net = tax_with_cess - total_tds_paid
    return RegimeResult(
        taxable_income=round(taxable_income),
        base_tax=round(base_tax),
        tax_with_cess=tax_with_cess,
        net_payable=round(abs(net)),
        is_refund=net < 0,
    )


@app.get("/")
def root():
    return {"status": "ok", "service": "ITR Tax Tracker API"}


@app.post("/calculate", response_model=TaxResult)
def calculate_tax(data: TaxInput):
    other_income_total = data.sb_interest + data.fd_interest + data.dividend_income
    total_tds_paid = data.tds_salary + data.tds_other + data.advance_tax

    # New regime: only standard deduction + employer NPS allowed
    new_salary_income = max(0.0, data.gross_salary - STANDARD_DEDUCTION_NEW - data.nps_employer)
    new_taxable_income = max(0.0, new_salary_income + other_income_total)
    new_base_tax, _ = compute_new_regime_tax(new_taxable_income, data.fy)
    new_result = build_regime_result(new_taxable_income, new_base_tax, total_tds_paid)

    # Old regime: exemptions + Chapter VI-A deductions + employer NPS all allowed
    total_exemptions = data.hra_exemption + data.lta_exemption + data.other_exemptions
    total_chapter_via = (
        data.ded_80c + data.ded_80ccd1b + data.ded_80d + data.ded_80tta_ttb + data.ded_other
    )
    old_salary_income = max(
        0.0, data.gross_salary - total_exemptions - STANDARD_DEDUCTION_OLD - data.nps_employer
    )
    old_taxable_income = max(0.0, old_salary_income + other_income_total - total_chapter_via)
    old_base_tax, _ = compute_old_regime_tax(old_taxable_income, data.age_group)
    old_result = build_regime_result(old_taxable_income, old_base_tax, total_tds_paid)

    if new_result.tax_with_cess < old_result.tax_with_cess:
        better = "new"
    elif old_result.tax_with_cess < new_result.tax_with_cess:
        better = "old"
    else:
        better = "same"

    return TaxResult(
        new_regime=new_result,
        old_regime=old_result,
        total_tds_paid=round(total_tds_paid),
        better_regime=better,
    )


# ---------------------------------------------------------------------------
# PDF EXTRACTION
# ---------------------------------------------------------------------------

def to_float(raw: str):
    try:
        return float(raw.replace(",", ""))
    except ValueError:
        return None


def _round_or_none(value):
    """Round to whole rupees, matching how amounts appear on the actual
    AIS/TIS/Form16/payslip documents — no decimal places."""
    return round(value) if value is not None else None


def extract_tis_category(text: str, category_name: str):
    """TIS has a clean summary table: 'Category Name   <processed>   <accepted>'.
    The 'accepted by taxpayer' figure (second number) is the authoritative one."""
    pattern = rf"{re.escape(category_name)}\s+([\d,]+)\s+([\d,]+)"
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        return to_float(match.group(2))
    return None


def sum_ais_lines_containing(text: str, marker: str):
    """AIS lists each bank/source as a separate line ending in the amount
    (e.g. 'SFT-016(SB) Interest income ... BANK NAME (CODE) 1 21,864').
    Sum the trailing amount across every matching line."""
    total = 0.0
    found = False
    for line in text.splitlines():
        if marker in line:
            numbers = re.findall(r"[\d,]+", line)
            if numbers:
                amount = to_float(numbers[-1])
                if amount is not None:
                    total += amount
                    found = True
    return total if found else None


def sum_salary_tds_deposited(text: str):
    """AIS salary section lists one row per quarter/payment:
    'SR QUARTER DATE AMOUNT_PAID TDS_DEDUCTED TDS_DEPOSITED STATUS'.
    Sum the TDS_DEPOSITED column across every such row."""
    total = 0.0
    found = False
    for line in text.splitlines():
        if re.match(r"^\d+\s+Q\d\(", line.strip()):
            numbers = re.findall(r"[\d,]+", line)
            if len(numbers) >= 4:
                tds_deposited = to_float(numbers[-1])
                if tds_deposited is not None:
                    total += tds_deposited
                    found = True
    return total if found else None


def extract_ais_gross_salary(text: str):
    """AIS Part B7 shows 'GROSS SALARY' as a distinct labeled figure, or the
    Part B1 salary row 'TDS-192 Salary received ... <count> <amount>'."""
    match = re.search(r"gross\s*salary\s*(?:received)?\D{0,20}?([\d,]+)", text, re.IGNORECASE)
    if match:
        return to_float(match.group(1))
    for line in text.splitlines():
        if "TDS-192" in line:
            numbers = re.findall(r"[\d,]+", line)
            if numbers:
                return to_float(numbers[-1])
    return None


GROSS_SALARY_PATTERNS = [
    r"gross\s*salary[^\d]{0,25}([\d,]+\.?\d*)",
    r"total\s*gross\s*(?:earnings|salary)[^\d]{0,25}([\d,]+\.?\d*)",
    r"gross\s*earnings[^\d]{0,25}([\d,]+\.?\d*)",
    r"total\s*earnings[^\d]{0,25}([\d,]+\.?\d*)",
]

TDS_PATTERNS = [
    r"total\s*tax\s*deducted[^\d]{0,25}([\d,]+\.?\d*)",
    r"tax\s*deducted\s*at\s*source[^\d]{0,25}([\d,]+\.?\d*)",
    r"total\s*(?:amount\s*of\s*)?tds[^\d]{0,25}([\d,]+\.?\d*)",
]


def extract_form16_figures(text: str):
    """Form 16 Part A has a reliable single-line summary:
    'Total (Rs.)  <amount paid/credited>  <TDS deducted>  <TDS deposited>'
    which is far more consistent across employers than trying to match
    'Gross Salary' labels (those often sit on a different line than their
    number, due to how the PDF's table columns get extracted)."""
    match = re.search(
        r"Total\s*\(Rs\.\)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)", text
    )
    if match:
        gross_salary = to_float(match.group(1))
        tds_deposited = to_float(match.group(3))
        return gross_salary, tds_deposited

    # Fallback: cross-line label matching for Form 16 layouts without that summary row
    gross_salary = None
    tds = None
    gross_match = re.search(
        r"Total amount of salary received from current employer.*?(\d[\d,]{3,}\.\d{2})",
        text, re.DOTALL,
    )
    if gross_match:
        gross_salary = to_float(gross_match.group(1))
    tds_match = re.search(
        r"Tax deducted from salary of the employee.*?(\d[\d,]{3,}\.\d{2})",
        text, re.DOTALL,
    )
    if tds_match:
        tds = to_float(tds_match.group(1))
    return gross_salary, tds


def find_amount(text: str, patterns: list[str]):
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return to_float(match.group(1))
    return None


def extract_pdf_text(file_bytes: bytes) -> str:
    text = ""
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text += (page.extract_text() or "") + "\n"
    return text


async def extract_pdf_text_async(file_bytes: bytes) -> str:
    """Run the CPU-bound pdfplumber parsing in a thread pool so multiple
    documents can be processed concurrently instead of one after another."""
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, extract_pdf_text, file_bytes)


@app.post("/extract", response_model=ExtractedFields)
async def extract_document(file: UploadFile = File(...), doc_type: str = Form(...)):
    """
    doc_type: "payslip" | "form16" | "ais" | "tis"

    TIS uses a clean summary table and is the preferred source for interest/
    dividend income. AIS lists the same figures spread across many repeating
    per-bank rows, which is summed as a fallback. Always verify extracted
    figures — layouts can still vary between years/portals.
    """
    file_bytes = await file.read()
    try:
        text = extract_pdf_text(file_bytes)
    except Exception as e:
        return ExtractedFields(doc_type=doc_type, raw_text_preview=f"Could not read PDF: {e}")

    result = ExtractedFields(doc_type=doc_type, raw_text_preview=text[:300])

    if doc_type == "payslip":
        result.gross_salary = _round_or_none(find_amount(text, GROSS_SALARY_PATTERNS))
        result.tds = _round_or_none(find_amount(text, TDS_PATTERNS))

    elif doc_type == "form16":
        gross_salary, tds = extract_form16_figures(text)
        result.gross_salary = _round_or_none(gross_salary)
        result.tds = _round_or_none(tds)

    elif doc_type == "tis":
        result.gross_salary = _round_or_none(extract_tis_category(text, "Salary"))
        result.sb_interest = _round_or_none(extract_tis_category(text, "Interest from savings bank"))
        result.fd_interest = _round_or_none(extract_tis_category(text, "Interest from deposit"))
        result.dividend_income = _round_or_none(extract_tis_category(text, "Dividend"))

    elif doc_type == "ais":
        result.gross_salary = _round_or_none(extract_ais_gross_salary(text))
        result.tds = _round_or_none(sum_salary_tds_deposited(text))
        result.sb_interest = _round_or_none(sum_ais_lines_containing(text, "SFT-016(SB)"))
        result.fd_interest = _round_or_none(sum_ais_lines_containing(text, "SFT-016(TD)"))
        result.dividend_income = _round_or_none(sum_ais_lines_containing(text, "Dividend"))

    return result


# ---------------------------------------------------------------------------
# FULL ANALYSIS: reconcile payslips vs Form 16, categorize every AIS/TIS
# income item per the Income Tax Act, compute salary + other-sources TDS
# separately, and estimate 234B/234C interest — following the same rounding
# conventions (Sec 288A / 288B, floor-to-nearest-100 before applying the
# monthly 234B/234C rate) used by professional ITR computation software, so
# the numbers reconcile exactly against a CA-prepared computation sheet.
# ---------------------------------------------------------------------------

# How each TIS category should be treated under the Income Tax Act.
# IMPORTANT: not everything reported in AIS/TIS is taxable income —
# "purchase of securities" and "outward remittance" are informational only.
TIS_CATEGORY_RULES = [
    (["salary"], "salary"),
    (["interest from savings bank"], "slab_other_income"),
    (["interest from deposit"], "slab_other_income"),
    (["dividend"], "slab_other_income"),
    (["receipts on transfer of virtual digital asset"], "vda_flat30"),
    (["purchase of securities", "purchase of immovable", "cash withdrawal", "cash deposit"], "non_income"),
    (["outward foreign remittance"], "non_income"),
    (["sale of securities", "sale of immovable", "capital gain"], "capital_gains_needs_review"),
]


def categorize_tis_category(name: str) -> str:
    name_lower = name.lower()
    for keywords, treatment in TIS_CATEGORY_RULES:
        if any(kw in name_lower for kw in keywords):
            return treatment
    return "needs_review"


def extract_tis_all_categories(text: str):
    """Parse every row of TIS's summary table (restricted to the summary
    section only — the detailed Annexure pages repeat similar-looking rows
    that would otherwise be mistakenly picked up)."""
    summary_section = text.split("Annexure to Taxpayer Information Summary")[0]
    pattern = r"^(\d+)\s+(.+?)\s+([\d,]+)\s+([\d,]+)$"
    rows = []
    for line in summary_section.splitlines():
        match = re.match(pattern, line.strip())
        if match:
            _, category, processed, accepted = match.groups()
            rows.append({"category": category.strip(), "amount": to_float(accepted)})
    return rows


def split_ais_salary_vs_other_tds(text: str):
    """AIS lists TDS as repeating quarterly rows under each category section.
    Split the document into the Salary section vs everything after it, and
    sum the 'TDS DEPOSITED' column separately for each."""
    salary_start = text.find("Salary\n")
    if salary_start == -1:
        return None, None

    # Find the next category header after Salary to bound the salary block.
    next_headers = ["Interest from deposit", "Interest from savings bank", "Receipts on transfer"]
    next_positions = [text.find(h, salary_start + 7) for h in next_headers]
    next_positions = [p for p in next_positions if p != -1]
    salary_end = min(next_positions) if next_positions else len(text)

    salary_block = text[salary_start:salary_end]
    rest_block = text[salary_end:]

    def sum_quarterly(block):
        total = 0.0
        found = False
        for line in block.splitlines():
            if re.match(r"^\d+\s+Q\d\(", line.strip()):
                numbers = re.findall(r"[\d,]+", line)
                if len(numbers) >= 4:
                    total += to_float(numbers[-1])
                    found = True
        return total if found else None

    return sum_quarterly(salary_block), sum_quarterly(rest_block)


def fy_ay_start_dates(fy_string: str):
    """Returns (fy_start, ay_start) for the given 'FY 2025-26' / 'FY 2024-25' string."""
    fy_start_year = 2025 if "2025-26" in fy_string else 2024
    fy_start = date(fy_start_year, 4, 1)
    ay_start = date(fy_start_year + 1, 4, 1)
    return fy_start, ay_start


def months_elapsed_since_fy_start(fy_string: str, today=None) -> int:
    """Months from 1 April of the Assessment Year to today — part of a month
    counts as a full month, per Section 234B."""
    today = today or datetime.now().date()
    _, ay_start = fy_ay_start_dates(fy_string)
    if today < ay_start:
        return 0
    months = (today.year - ay_start.year) * 12 + (today.month - ay_start.month)
    if today.day > ay_start.day:
        months += 1
    return max(0, months)


def add_months(d: date, n: int) -> date:
    month_index = d.month - 1 + n
    year = d.year + month_index // 12
    month = month_index % 12 + 1
    return date(year, month, 1)


def compute_234b_detailed(assessed_tax: float, fy: str, today=None):
    """Section 234B, computed exactly as ITR software does it: the assessed
    tax (total tax minus TDS minus advance tax already paid) is floored to
    the nearest Rs 100 BEFORE the 1%/month rate is applied. If assessed tax
    is <= Rs 10,000, there's no advance-tax obligation at all (Section 208),
    so no interest applies either."""
    months = months_elapsed_since_fy_start(fy, today)
    if assessed_tax <= SECTION_208_THRESHOLD or months <= 0:
        return 0, months, []

    principal_rounded = floor_100(assessed_tax)
    monthly_interest = round_half_up(principal_rounded * 0.01)
    _, ay_start = fy_ay_start_dates(fy)
    breakdown = []
    total_interest = 0
    for i in range(months):
        m = add_months(ay_start, i)
        breakdown.append({
            "month": m.strftime("%B-%Y"),
            "principal": round(assessed_tax),
            "interest": monthly_interest,
        })
        total_interest += monthly_interest
    return total_interest, months, breakdown


MONTH_NAMES = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _parse_date_token(part1: str, part2: str, part3: str):
    try:
        if part2.isalpha():
            month = MONTH_NAMES.get(part2.lower()[:3])
            if not month:
                return None
            return date(int(part3), month, int(part1))
        d, m, y = int(part1), int(part2), int(part3)
        return date(y, m, d)
    except (ValueError, TypeError):
        return None


def extract_vda_transfer_date(text: str, fy: str):
    """Best-effort, fully automatic: locates the 'Receipts on transfer of
    virtual digital asset' block in AIS text and returns the LATEST
    transaction date found within it (bounded to that section only, so
    dates from an unrelated section that follows - e.g. outward foreign
    remittance - never leak in).

    AIS reports each VDA sale as a separate dated row (e.g. 31/05/2025 and
    16/06/2025 for two disposals that together make up one reported total).
    Real computation software treats the whole reported amount as a single
    transaction dated at the LAST of those rows rather than prorating tax
    across each one - using the latest date here reproduces that same
    convention, so 234C matches a CA-prepared computation sheet. Returns
    None if nothing confidently found, in which case 234C falls back to the
    safe full-year simplification rather than guessing."""
    if not text:
        return None
    fy_start, ay_start = fy_ay_start_dates(fy)
    fy_end = date(ay_start.year, 3, 31)

    lower = text.lower()
    idx = lower.find("receipts on transfer of virtual digital asset")
    if idx == -1:
        idx = lower.find("virtual digital asset")
    if idx == -1:
        return None

    # Bound the block to just this category: keep consuming lines that are
    # either row/label continuations (start with a digit, "SR." or "(") -
    # stop at the first line that looks like the start of the NEXT category
    # header, so we never wander into an unrelated section.
    lines = text[idx:].splitlines()
    block_lines = []
    for i, line in enumerate(lines):
        stripped = line.strip()
        if i > 0 and stripped and not re.match(r"^(SR\.?\s*NO\.?|\d|\()", stripped, re.IGNORECASE):
            break
        block_lines.append(line)
        if len(block_lines) > 40:
            break
    block = "\n".join(block_lines)

    candidates = []
    # Prefer the precise per-transaction row format used throughout AIS:
    # "<sr> Q#(Mon-Mon) DD/MM/YYYY <amount paid/credited> <tds> <tds> <status>"
    for line in block.splitlines():
        if re.match(r"^\d+\s+Q\d\(", line.strip()):
            m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", line)
            if m:
                c = _parse_date_token(m.group(1), m.group(2), m.group(3))
                if c and fy_start <= c <= fy_end:
                    candidates.append(c)
    if not candidates:
        # Fallback: any date-like token anywhere in the bounded block.
        for m in re.finditer(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", block):
            c = _parse_date_token(m.group(1), m.group(2), m.group(3))
            if c and fy_start <= c <= fy_end:
                candidates.append(c)

    return max(candidates) if candidates else None


def compute_234c_detailed(
    slab_only_tax_with_cess: float,
    full_tax_with_cess: float,
    total_tds: float,
    advance_tax_paid: float,
    fy: str,
    vda_transfer_date: date | None = None,
    today=None,
):
    """Section 234C, computed against four installment checkpoints
    (15%/45%/75%/100% by 15 Jun / 15 Sep / 15 Dec / 15 Mar). Same floor-to-
    Rs-100 rounding as 234B is applied to each checkpoint's shortfall before
    the rate is applied.

    Two automatic (no manual input) refinements are layered on top of the
    plain calculation:
    1. Any checkpoint whose due date hasn't occurred yet as of today is
       skipped entirely - there's no advance-tax shortfall for an
       installment that hasn't come due yet (this is what makes 234B/234C
       "calculate only up to whichever month you're checking in").
    2. If `vda_transfer_date` was auto-extracted from the uploaded AIS/TIS
       (see extract_vda_transfer_date), any checkpoint due BEFORE that date
       excludes the VDA/capital-gains tax component - you can't owe advance
       tax on a gain that hadn't happened yet. When no date could be found,
       every checkpoint uses the full assessed tax (the same simplified
       "from start of year" mode professional ITR software itself offers as
       a toggle) - a small, always-conservative overstatement rather than a
       guess."""
    today = today or datetime.now().date()
    full_assessed = full_tax_with_cess - total_tds
    if full_assessed <= SECTION_208_THRESHOLD:
        return 0, []

    slab_only_assessed = slab_only_tax_with_cess - total_tds
    fy_start_year = 2025 if "2025-26" in fy else 2024
    checkpoints = [
        (0.15, 3, "First (Up to 15 Jun)", date(fy_start_year, 6, 15)),
        (0.45, 3, "Second (Up to 15 Sep)", date(fy_start_year, 9, 15)),
        (0.75, 3, "Third (Up to 15 Dec)", date(fy_start_year, 12, 15)),
        (1.00, 1, "Fourth (Up to 15 Mar)", date(fy_start_year + 1, 3, 15)),
    ]

    breakdown = []
    total_interest = 0
    for pct, months, label, due_date in checkpoints:
        if due_date > today:
            continue  # installment hasn't come due yet - no shortfall possible

        if vda_transfer_date is not None and vda_transfer_date > due_date:
            assessed_for_checkpoint = max(0.0, slab_only_assessed)
        else:
            assessed_for_checkpoint = max(0.0, full_assessed)

        required = round_half_up(assessed_for_checkpoint * pct)
        shortfall = max(0, required - advance_tax_paid)
        shortfall_rounded = floor_100(shortfall)
        interest = round_half_up(shortfall_rounded * 0.01 * months)
        total_interest += interest
        breakdown.append({
            "installment": label,
            "required_pct": pct * 100,
            "required_amount": required,
            "remaining_due_rounded": shortfall_rounded,
            "months": months,
            "interest": interest,
        })
    return total_interest, breakdown


class SlabBreakdownItem(BaseModel):
    from_amount: int
    to_amount: int
    rate_pct: float
    taxable_amount: int
    tax: int


class InterestMonthItem(BaseModel):
    month: str
    principal: int
    interest: int


class InterestCheckpointItem(BaseModel):
    installment: str
    required_pct: float
    required_amount: int
    remaining_due_rounded: int
    months: int
    interest: int


class IncomeItem(BaseModel):
    category: str
    amount: int
    treatment: str  # slab_other_income | vda_flat30 | non_income | capital_gains_needs_review | needs_review
    included_in_tax: bool


class SalaryReconciliation(BaseModel):
    payslip_total: int | None
    form16_total: int | None
    additional_income_identified: int
    final_gross_salary: int


class TDSReconciliation(BaseModel):
    payslip_tds_total: int | None
    form16_tds: int | None
    salary_tds_used: int
    other_sources_tds: int
    total_tds: int


class RegimeFullResult(BaseModel):
    salary_income: int
    gross_total_income: int
    total_income_rounded_288a: int
    normal_income: int
    slab_tax_breakdown: list[SlabBreakdownItem]
    slab_tax: int
    special_rate_income: int
    special_rate_tax: int
    base_tax: int
    cess: int
    taxable_income: int
    slab_tax_with_cess: int
    vda_tax_with_cess: int
    total_tax: int
    total_tds_paid: int
    advance_tax_paid: int
    net_before_interest: int
    interest_234b: int
    interest_234c: int
    months_elapsed_234b: int
    interest_234b_breakdown: list[InterestMonthItem]
    interest_234c_breakdown: list[InterestCheckpointItem]
    amount_before_288b: int
    rounded_288b: int
    self_assessment_tax_paid: int
    final_amount: int
    is_refund: bool


class FullAnalysisResult(BaseModel):
    salary: SalaryReconciliation
    tds: TDSReconciliation
    other_income_items: list[IncomeItem]
    vda_income_total: int
    slab_other_income_total: int
    non_income_total: int
    new_regime: RegimeFullResult
    old_regime: RegimeFullResult
    vda_caveat: str
    calculation_date: str
    itr_due_date_note: str


async def run_full_analysis(
    fy: str,
    age_group: str,
    nps_employer: float,
    hra_exemption: float,
    lta_exemption: float,
    other_exemptions: float,
    ded_80c: float,
    ded_80ccd1b: float,
    ded_80d: float,
    ded_80tta_ttb: float,
    ded_other: float,
    advance_tax: float,
    self_assessment_tax_paid: float,
    additional_vda_income: float,
    payslips: list[UploadFile],
    form16: UploadFile | None,
    ais: UploadFile | None,
    tis: UploadFile | None,
) -> FullAnalysisResult:
    """Shared computation core used by both /full-analysis (JSON for the UI)
    and /export-excel (the downloadable workbook) — kept as one function so
    the two can never drift apart."""
    # --- Read and extract ALL documents concurrently (instead of one after
    # another) — this is the main lever we control for speed, since the PDF
    # parsing itself is CPU-bound but independent per document. ---
    payslip_bytes = [await f.read() for f in payslips]
    form16_bytes = await form16.read() if form16 else None
    ais_bytes = await ais.read() if ais else None
    tis_bytes = await tis.read() if tis else None

    extraction_tasks = [extract_pdf_text_async(b) for b in payslip_bytes]
    if form16_bytes is not None:
        extraction_tasks.append(extract_pdf_text_async(form16_bytes))
    if ais_bytes is not None:
        extraction_tasks.append(extract_pdf_text_async(ais_bytes))
    if tis_bytes is not None:
        extraction_tasks.append(extract_pdf_text_async(tis_bytes))

    extracted_texts = await asyncio.gather(*extraction_tasks) if extraction_tasks else []

    idx = 0
    payslip_texts = extracted_texts[idx: idx + len(payslip_bytes)]
    idx += len(payslip_bytes)
    form16_text = extracted_texts[idx] if form16_bytes is not None else None
    idx += 1 if form16_bytes is not None else 0
    ais_text = extracted_texts[idx] if ais_bytes is not None else ""
    idx += 1 if ais_bytes is not None else 0
    tis_text = extracted_texts[idx] if tis_bytes is not None else None

    # --- Salary reconciliation: payslips vs Form 16 ---
    payslip_total = None
    payslip_tds_total = None
    if payslip_texts:
        total_gross, total_tds, found = 0.0, 0.0, False
        for text in payslip_texts:
            g = find_amount(text, GROSS_SALARY_PATTERNS)
            t = find_amount(text, TDS_PATTERNS)
            if g:
                total_gross += g
                found = True
            if t:
                total_tds += t
        payslip_total = round(total_gross) if found else None
        payslip_tds_total = round(total_tds) if total_tds else None

    form16_gross, form16_tds = None, None
    if form16_text is not None:
        form16_gross, form16_tds = extract_form16_figures(form16_text)
        form16_gross = round(form16_gross) if form16_gross else None
        form16_tds = round(form16_tds) if form16_tds else None

    # Form 16 is authoritative; any excess over the payslip total (e.g.
    # perquisites, bonuses not reflected on payslips) is surfaced explicitly.
    if form16_gross is not None:
        base = payslip_total or 0
        additional = max(0, form16_gross - base)
        final_gross_salary = base + additional
    elif payslip_total is not None:
        additional = 0
        final_gross_salary = payslip_total
    else:
        additional = 0
        final_gross_salary = 0

    salary_recon = SalaryReconciliation(
        payslip_total=payslip_total,
        form16_total=form16_gross,
        additional_income_identified=additional,
        final_gross_salary=final_gross_salary,
    )

    # --- AIS: split salary TDS vs other-sources TDS ---
    ais_salary_tds, ais_other_tds = None, None
    if ais_text:
        ais_salary_tds, ais_other_tds = split_ais_salary_vs_other_tds(ais_text)

    salary_tds_used = form16_tds or ais_salary_tds or payslip_tds_total or 0
    other_sources_tds = round(ais_other_tds) if ais_other_tds else 0

    tds_recon = TDSReconciliation(
        payslip_tds_total=payslip_tds_total,
        form16_tds=form16_tds,
        salary_tds_used=round(salary_tds_used),
        other_sources_tds=other_sources_tds,
        total_tds=round(salary_tds_used) + other_sources_tds,
    )

    # --- TIS: categorize every income item per the Income Tax Act ---
    other_income_items = []
    slab_other_income_total = 0.0
    vda_income_total = 0.0
    non_income_total = 0.0

    if tis_text:
        for row in extract_tis_all_categories(tis_text):
            if row["category"].lower() == "salary" or row["amount"] is None:
                continue
            treatment = categorize_tis_category(row["category"])
            included = treatment in ("slab_other_income", "vda_flat30")
            other_income_items.append(IncomeItem(
                category=row["category"], amount=round(row["amount"]),
                treatment=treatment, included_in_tax=included,
            ))
            if treatment == "slab_other_income":
                slab_other_income_total += row["amount"]
            elif treatment == "vda_flat30":
                vda_income_total += row["amount"]
            elif treatment == "non_income":
                non_income_total += row["amount"]

    # Manual top-up for VDA / capital gains income not picked up from TIS
    # (e.g. TIS summary table doesn't carry per-transaction dates).
    vda_income_total += additional_vda_income

    # Fully automatic - no manual date entry: try to pull the VDA transfer
    # date straight out of the uploaded AIS (preferred, has per-transaction
    # detail) or TIS text. If nothing is confidently found, 234C falls back
    # to the safe full-year simplification further down.
    auto_vda_transfer_date = (
        extract_vda_transfer_date(ais_text, fy)
        or extract_vda_transfer_date(tis_text or "", fy)
    )

    # --- Tax calculation, both regimes ---
    def calc_regime(is_new_regime: bool) -> RegimeFullResult:
        if is_new_regime:
            salary_income = max(0.0, final_gross_salary - STANDARD_DEDUCTION_NEW - nps_employer)
            normal_income_unrounded = max(0.0, salary_income + slab_other_income_total)
        else:
            total_exemptions = hra_exemption + lta_exemption + other_exemptions
            total_via = ded_80c + ded_80ccd1b + ded_80d + ded_80tta_ttb + ded_other
            salary_income = max(0.0, final_gross_salary - total_exemptions - STANDARD_DEDUCTION_OLD - nps_employer)
            normal_income_unrounded = max(0.0, salary_income + slab_other_income_total - total_via)

        gross_total_income = normal_income_unrounded + vda_income_total
        total_income_rounded_288a = round_10(gross_total_income)

        # Section 288A rounding is applied to the "normal" (slab-rate) income
        # BEFORE slab tax is computed — special-rate (VDA) income is taxed
        # separately on its unrounded value.
        normal_income_rounded = round_10(normal_income_unrounded)
        if is_new_regime:
            slab_tax_raw, slab_breakdown_raw = compute_new_regime_tax(normal_income_rounded, fy)
        else:
            slab_tax_raw, slab_breakdown_raw = compute_old_regime_tax(normal_income_rounded, age_group)

        slab_tax = round(slab_tax_raw)
        slab_breakdown = [
            SlabBreakdownItem(
                from_amount=b["from"], to_amount=b["to"], rate_pct=b["rate_pct"],
                taxable_amount=b["taxable_amount"], tax=b["tax"],
            ) for b in slab_breakdown_raw
        ]

        # VDA / capital gains: flat 30% + cess, no deductions, Sec 115BBH —
        # same treatment in both regimes.
        special_rate_tax = round_half_up(vda_income_total * 0.30)

        base_tax = slab_tax + special_rate_tax
        cess = round_half_up(base_tax * 0.04)
        total_tax_with_cess = base_tax + cess

        # Kept for backward compatibility / simpler summary display.
        slab_tax_with_cess = round_half_up(slab_tax * 1.04)
        vda_tax_with_cess = round_half_up(special_rate_tax * 1.04)

        total_tds = tds_recon.total_tds
        net_before_interest = total_tax_with_cess - total_tds - advance_tax

        assessed_tax = total_tax_with_cess - total_tds - advance_tax
        if assessed_tax > 0:
            interest_b, months_b, breakdown_b = compute_234b_detailed(assessed_tax, fy)
            interest_c, breakdown_c = compute_234c_detailed(
                slab_tax_with_cess, total_tax_with_cess, total_tds, advance_tax, fy,
                vda_transfer_date=auto_vda_transfer_date,
            )
        else:
            interest_b, months_b, breakdown_b = 0, months_elapsed_since_fy_start(fy), []
            interest_c, breakdown_c = 0, []

        amount_before_288b = net_before_interest + interest_b + interest_c
        rounded_288b = round_10(amount_before_288b)
        final_after_self_assessment = rounded_288b - self_assessment_tax_paid

        return RegimeFullResult(
            salary_income=round(salary_income),
            gross_total_income=round(gross_total_income),
            total_income_rounded_288a=total_income_rounded_288a,
            normal_income=round(normal_income_rounded),
            slab_tax_breakdown=slab_breakdown,
            slab_tax=slab_tax,
            special_rate_income=round(vda_income_total),
            special_rate_tax=special_rate_tax,
            base_tax=base_tax,
            cess=cess,
            taxable_income=round(normal_income_rounded + vda_income_total),
            slab_tax_with_cess=slab_tax_with_cess,
            vda_tax_with_cess=vda_tax_with_cess,
            total_tax=total_tax_with_cess,
            total_tds_paid=round(total_tds),
            advance_tax_paid=round(advance_tax),
            net_before_interest=round(net_before_interest),
            interest_234b=interest_b,
            interest_234c=interest_c,
            months_elapsed_234b=months_b,
            interest_234b_breakdown=[InterestMonthItem(**m) for m in breakdown_b],
            interest_234c_breakdown=[InterestCheckpointItem(**c) for c in breakdown_c],
            amount_before_288b=round(amount_before_288b),
            rounded_288b=rounded_288b,
            self_assessment_tax_paid=round(self_assessment_tax_paid),
            final_amount=round(abs(final_after_self_assessment)),
            is_refund=final_after_self_assessment < 0,
        )

    today = datetime.now().date()
    _, ay_start = fy_ay_start_dates(fy)
    itr_due_date = date(ay_start.year, 7, 31)
    if today <= itr_due_date:
        due_note = f"Today ({today.strftime('%d %b %Y')}) is before the {itr_due_date.strftime('%d %b %Y')} ITR filing deadline — 234B is calculated only for the months actually elapsed so far, not the full year."
    else:
        due_note = f"Today ({today.strftime('%d %b %Y')}) is after the {itr_due_date.strftime('%d %b %Y')} ITR filing deadline — file as soon as possible, additional interest under Section 234A may also apply for late filing (not calculated here)."

    return FullAnalysisResult(
        salary=salary_recon,
        tds=tds_recon,
        other_income_items=other_income_items,
        vda_income_total=round(vda_income_total),
        slab_other_income_total=round(slab_other_income_total),
        non_income_total=round(non_income_total),
        new_regime=calc_regime(True),
        old_regime=calc_regime(False),
        vda_caveat=(
            "AIS/TIS report the gross transaction value for virtual digital "
            "assets, not your actual gain or loss. The amount above may "
            "overstate your real taxable VDA income if your cost basis was "
            "high, or if this included transfers between your own wallets. "
            "Verify against your exchange statements before relying on this."
        ),
        calculation_date=today.strftime("%d %b %Y"),
        itr_due_date_note=due_note,
    )


# ---------------------------------------------------------------------------
# EXCEL EXPORT: a professionally formatted, multi-sheet workbook mirroring
# the Detailed Computation view - Summary, Income & TDS Detail, and one
# fully-worked sheet per regime.
# ---------------------------------------------------------------------------

INR_FMT = '#,##,##0;[RED]-#,##,##0'
XL_PRIMARY = "0E7C6B"
XL_PRIMARY_DARK = "0A5C4F"
XL_GOLD = "B7791F"
XL_LIGHT_BG = "F4F6F8"
XL_INK = "1A2027"
XL_INK_MUTED = "6B7280"
XL_BORDER = "D9DCE1"
XL_WHITE = "FFFFFF"

_thin_side = Side(style="thin", color=XL_BORDER)
_TABLE_BORDER = Border(left=_thin_side, right=_thin_side, top=_thin_side, bottom=_thin_side)


def _sheet_title(ws, title, subtitle, n_cols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(name="Calibri", size=16, bold=True, color=XL_WHITE)
    cell.fill = PatternFill("solid", fgColor=XL_PRIMARY_DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub = ws.cell(row=2, column=1, value=subtitle)
    sub.font = Font(name="Calibri", size=10, italic=True, color=XL_WHITE)
    sub.fill = PatternFill("solid", fgColor=XL_PRIMARY)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20
    return 4


def _section_header(ws, row, n_cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=11, bold=True, color=XL_WHITE)
    cell.fill = PatternFill("solid", fgColor=XL_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 20
    return row + 1


def _kv_row(ws, row, label, value, n_cols, bold=False, indent=0, fmt=INR_FMT, is_total=False, muted=False):
    label_cell = ws.cell(row=row, column=1, value=("   " * indent) + label)
    label_cell.font = Font(name="Calibri", size=10, bold=bold, italic=muted,
                            color=XL_INK_MUTED if muted else XL_INK)
    if n_cols > 2:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n_cols - 1)
    value_cell = ws.cell(row=row, column=n_cols, value=value)
    value_cell.font = Font(name="Calibri", size=10, bold=bold, color=XL_INK)
    if isinstance(value, (int, float)):
        value_cell.number_format = fmt
    value_cell.alignment = Alignment(horizontal="right")
    if is_total:
        top = Side(style="thin", color=XL_INK_MUTED)
        label_cell.border = Border(top=top)
        value_cell.border = Border(top=top)
    return row + 1


def _data_table(ws, row, headers, rows, col_widths=None):
    """Writes a bordered, header-styled table starting at `row`, returns the
    next free row. `rows` is a list of tuples matching `headers` length."""
    n_cols = len(headers)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Calibri", size=9, bold=True, color=XL_WHITE)
        cell.fill = PatternFill("solid", fgColor=XL_PRIMARY_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _TABLE_BORDER
    row += 1
    for r_i, data_row in enumerate(rows):
        for c, val in enumerate(data_row, start=1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = _TABLE_BORDER
            cell.font = Font(name="Calibri", size=9.5, color=XL_INK)
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = INR_FMT
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
            if r_i % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=XL_LIGHT_BG)
        row += 1
    return row


def _write_regime_sheet(ws, result: FullAnalysisResult, regime: RegimeFullResult, label: str, fy: str):
    n_cols = 4
    row = _sheet_title(ws, f"Detailed Computation - {label}", f"{fy} | Calculated as of {result.calculation_date}", n_cols)

    row = _section_header(ws, row, n_cols, "Computation of Total Income")
    std_ded = STANDARD_DEDUCTION_NEW if "New" in label else STANDARD_DEDUCTION_OLD
    row = _kv_row(ws, row, "Gross Salary", result.salary.final_gross_salary, n_cols, indent=1, muted=True)
    row = _kv_row(ws, row, "Less: Standard Deduction", -std_ded, n_cols, indent=1, muted=True)
    row = _kv_row(ws, row, "Income from Salary", regime.salary_income, n_cols, bold=True)
    row = _kv_row(ws, row, "Income from Other Sources", result.slab_other_income_total, n_cols)
    if regime.special_rate_income > 0:
        row = _kv_row(ws, row, "Income from Capital Gains (VDA - Sec 115BBH)", regime.special_rate_income, n_cols)
    row = _kv_row(ws, row, "Gross Total Income", regime.gross_total_income, n_cols, bold=True, is_total=True)
    row = _kv_row(ws, row, "Round off u/s 288A", regime.total_income_rounded_288a, n_cols)
    row += 1

    row = _section_header(ws, row, n_cols, f"Tax on Normal Income (Rs {regime.normal_income:,})")
    if regime.slab_tax_breakdown:
        row = _data_table(
            ws, row, ["Slab From", "Slab To", "Rate", "Tax"],
            [(b.from_amount, b.to_amount, f"{b.rate_pct}%", b.tax) for b in regime.slab_tax_breakdown],
        )
    else:
        row = _kv_row(ws, row, "No slab tax - within rebate/exemption limit", "", n_cols, muted=True)
    row = _kv_row(ws, row, "Tax on Normal Income", regime.slab_tax, n_cols, bold=True, is_total=True)
    if regime.special_rate_income > 0:
        row = _kv_row(ws, row, "Tax on Special Rate Income (VDA @ 30%)", regime.special_rate_tax, n_cols)
    row = _kv_row(ws, row, "Total Tax", regime.base_tax, n_cols, bold=True, is_total=True)
    row = _kv_row(ws, row, "Health & Education Cess @ 4%", regime.cess, n_cols)
    row = _kv_row(ws, row, "Total Tax + Cess", regime.total_tax, n_cols, bold=True, is_total=True)
    row += 1

    row = _section_header(ws, row, n_cols, "Prepaid Taxes")
    row = _kv_row(ws, row, "T.D.S. - Salary", -result.tds.salary_tds_used, n_cols)
    row = _kv_row(ws, row, "T.D.S. - Non-Salary", -result.tds.other_sources_tds, n_cols)
    if regime.advance_tax_paid > 0:
        row = _kv_row(ws, row, "Advance Tax Paid", -regime.advance_tax_paid, n_cols)
    row = _kv_row(ws, row, "Balance (before interest)", regime.net_before_interest, n_cols, bold=True, is_total=True)
    row += 1

    if regime.interest_234b > 0:
        row = _section_header(ws, row, n_cols, "Interest Calculation u/s 234B")
        row = _data_table(
            ws, row, ["Month", "Principal", "Interest @ 1%"],
            [(m.month, m.principal, m.interest) for m in regime.interest_234b_breakdown],
        )
        row = _kv_row(ws, row, "Total Interest u/s 234B", regime.interest_234b, n_cols, bold=True, is_total=True)
        row += 1

    if regime.interest_234c > 0:
        row = _section_header(ws, row, n_cols, "Interest Calculation u/s 234C")
        row = _data_table(
            ws, row, ["Installment", "Required %", "Required Amt", "Remaining Due (rounded)", "Interest"],
            [(c.installment, f"{c.required_pct}%", c.required_amount, c.remaining_due_rounded, c.interest)
             for c in regime.interest_234c_breakdown],
        )
        row = _kv_row(ws, row, "Total Interest u/s 234C", regime.interest_234c, n_cols, bold=True, is_total=True)
        row += 1

    row = _section_header(ws, row, n_cols, "Final Settlement")
    row = _kv_row(ws, row, "Balance + Interest (234B + 234C)", regime.amount_before_288b, n_cols)
    row = _kv_row(ws, row, "Round off u/s 288B", regime.rounded_288b, n_cols)
    if regime.self_assessment_tax_paid > 0:
        row = _kv_row(ws, row, "Less: Self-Assessment Tax Deposited (u/s 140A)", -regime.self_assessment_tax_paid, n_cols)
    final_label = "Refund Due" if regime.is_refund else "Tax Payable"
    final_cell_row = row
    row = _kv_row(ws, row, final_label, regime.final_amount, n_cols, bold=True, is_total=True)
    for c in range(1, n_cols + 1):
        ws.cell(row=final_cell_row, column=c).fill = PatternFill(
            "solid", fgColor="FCF3E3" if not regime.is_refund else "E6F4F1"
        )

    ws.column_dimensions["A"].width = 42
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 16
    ws.sheet_view.showGridLines = False


def build_excel_workbook(result: FullAnalysisResult, fy: str, age_group: str) -> io.BytesIO:
    wb = Workbook()

    # --- Summary sheet ---
    ws = wb.active
    ws.title = "Summary"
    n_cols = 3
    row = _sheet_title(ws, "Income Tax Computation Summary", f"{fy} | Age category: {age_group} | Calculated as of {result.calculation_date}", n_cols)

    headers_row = row
    ws.cell(row=row, column=1, value="").fill = PatternFill("solid", fgColor=XL_PRIMARY_DARK)
    for c, h in enumerate(["Figure", "New Regime", "Old Regime"], start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = Font(name="Calibri", size=10, bold=True, color=XL_WHITE)
        cell.fill = PatternFill("solid", fgColor=XL_PRIMARY_DARK)
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left", indent=1 if c == 1 else 0)
        cell.border = _TABLE_BORDER
    row += 1

    def summary_row(label, new_val, old_val, bold=False, is_total=False):
        nonlocal row
        cells = [
            (1, label, "left"),
            (2, new_val, "right"),
            (3, old_val, "right"),
        ]
        for c, val, align in cells:
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = Font(name="Calibri", size=10, bold=bold, color=XL_INK)
            cell.alignment = Alignment(horizontal=align)
            cell.border = _TABLE_BORDER
            if isinstance(val, (int, float)):
                cell.number_format = INR_FMT
            if row % 2 == 1:
                cell.fill = PatternFill("solid", fgColor=XL_LIGHT_BG)
        row += 1

    summary_row("Gross Total Income", result.new_regime.gross_total_income, result.old_regime.gross_total_income)
    summary_row("Total Income (rounded u/s 288A)", result.new_regime.total_income_rounded_288a, result.old_regime.total_income_rounded_288a)
    summary_row("Tax on Normal Income", result.new_regime.slab_tax, result.old_regime.slab_tax)
    summary_row("Tax on Special Rate Income (VDA)", result.new_regime.special_rate_tax, result.old_regime.special_rate_tax)
    summary_row("Total Tax + Cess", result.new_regime.total_tax, result.old_regime.total_tax, bold=True)
    summary_row("Total TDS Paid", result.new_regime.total_tds_paid, result.old_regime.total_tds_paid)
    summary_row("Interest u/s 234B", result.new_regime.interest_234b, result.old_regime.interest_234b)
    summary_row("Interest u/s 234C", result.new_regime.interest_234c, result.old_regime.interest_234c)
    summary_row("Self-Assessment Tax Deposited", result.new_regime.self_assessment_tax_paid, result.old_regime.self_assessment_tax_paid)
    summary_row(
        "FINAL AMOUNT (Payable / Refund)",
        (-result.new_regime.final_amount if result.new_regime.is_refund else result.new_regime.final_amount),
        (-result.old_regime.final_amount if result.old_regime.is_refund else result.old_regime.final_amount),
        bold=True, is_total=True,
    )
    row += 1
    note = ws.cell(row=row, column=1, value="Negative figures in the final row indicate a refund due to you, not an amount payable.")
    note.font = Font(name="Calibri", size=9, italic=True, color=XL_INK_MUTED)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.sheet_view.showGridLines = False

    # --- Income & TDS detail sheet ---
    ws2 = wb.create_sheet("Income & TDS Detail")
    row = _sheet_title(ws2, "Income & TDS Reconciliation", f"{fy} | Calculated as of {result.calculation_date}", 3)
    row = _section_header(ws2, row, 3, "Salary Reconciliation")
    row = _kv_row(ws2, row, "Payslip total (summed)", result.salary.payslip_total or 0, 3)
    row = _kv_row(ws2, row, "Form 16 gross salary", result.salary.form16_total or 0, 3)
    row = _kv_row(ws2, row, "Additional income identified & added", result.salary.additional_income_identified, 3)
    row = _kv_row(ws2, row, "Final gross salary used", result.salary.final_gross_salary, 3, bold=True, is_total=True)
    row += 1
    row = _section_header(ws2, row, 3, "TDS Reconciliation")
    row = _kv_row(ws2, row, "Salary TDS used", result.tds.salary_tds_used, 3)
    row = _kv_row(ws2, row, "TDS on other income (AIS)", result.tds.other_sources_tds, 3)
    row = _kv_row(ws2, row, "Total TDS paid", result.tds.total_tds, 3, bold=True, is_total=True)
    row += 1
    if result.other_income_items:
        row = _section_header(ws2, row, 3, "Other Income - Categorized per Income Tax Act")
        row = _data_table(
            ws2, row, ["Category", "Amount", "Treatment"],
            [(i.category, i.amount, i.treatment.replace("_", " ").title()) for i in result.other_income_items],
        )
    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 22
    ws2.sheet_view.showGridLines = False

    # --- One detailed sheet per regime ---
    ws_new = wb.create_sheet("New Regime")
    _write_regime_sheet(ws_new, result, result.new_regime, "New Regime (Sec 115BAC)", fy)

    ws_old = wb.create_sheet("Old Regime")
    _write_regime_sheet(ws_old, result, result.old_regime, "Old Regime", fy)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


@app.post("/full-analysis", response_model=FullAnalysisResult)
async def full_analysis(
    fy: str = Form(...),
    age_group: str = Form("Below 60"),
    nps_employer: float = Form(0.0),
    hra_exemption: float = Form(0.0),
    lta_exemption: float = Form(0.0),
    other_exemptions: float = Form(0.0),
    ded_80c: float = Form(0.0),
    ded_80ccd1b: float = Form(0.0),
    ded_80d: float = Form(0.0),
    ded_80tta_ttb: float = Form(0.0),
    ded_other: float = Form(0.0),
    advance_tax: float = Form(0.0),
    self_assessment_tax_paid: float = Form(0.0),
    additional_vda_income: float = Form(0.0),
    payslips: list[UploadFile] = File(default=[]),
    form16: UploadFile | None = File(default=None),
    ais: UploadFile | None = File(default=None),
    tis: UploadFile | None = File(default=None),
):
    return await run_full_analysis(
        fy, age_group, nps_employer, hra_exemption, lta_exemption, other_exemptions,
        ded_80c, ded_80ccd1b, ded_80d, ded_80tta_ttb, ded_other,
        advance_tax, self_assessment_tax_paid, additional_vda_income,
        payslips, form16, ais, tis,
    )


@app.post("/export-excel")
async def export_excel(
    fy: str = Form(...),
    age_group: str = Form("Below 60"),
    nps_employer: float = Form(0.0),
    hra_exemption: float = Form(0.0),
    lta_exemption: float = Form(0.0),
    other_exemptions: float = Form(0.0),
    ded_80c: float = Form(0.0),
    ded_80ccd1b: float = Form(0.0),
    ded_80d: float = Form(0.0),
    ded_80tta_ttb: float = Form(0.0),
    ded_other: float = Form(0.0),
    advance_tax: float = Form(0.0),
    self_assessment_tax_paid: float = Form(0.0),
    additional_vda_income: float = Form(0.0),
    payslips: list[UploadFile] = File(default=[]),
    form16: UploadFile | None = File(default=None),
    ais: UploadFile | None = File(default=None),
    tis: UploadFile | None = File(default=None),
):
    result = await run_full_analysis(
        fy, age_group, nps_employer, hra_exemption, lta_exemption, other_exemptions,
        ded_80c, ded_80ccd1b, ded_80d, ded_80tta_ttb, ded_other,
        advance_tax, self_assessment_tax_paid, additional_vda_income,
        payslips, form16, ais, tis,
    )
    workbook_bytes = build_excel_workbook(result, fy, age_group)
    filename = f"Tax_Computation_{fy.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        workbook_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
